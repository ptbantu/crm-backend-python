#!/usr/bin/env python3
"""
分析 Accounts.xlsx 文件结构
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from openpyxl import load_workbook
    
    file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'docs/excel/Accounts.xlsx')
    wb = load_workbook(file_path)
    ws = wb.active
    
    print("=" * 80)
    print("Accounts.xlsx 文件分析")
    print("=" * 80)
    print(f"\n工作表名称: {ws.title}")
    print(f"最大行数: {ws.max_row}")
    print(f"最大列数: {ws.max_column}")
    
    # 读取标题行
    print("\n" + "=" * 80)
    print("列标题:")
    print("=" * 80)
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
        print(f"{cell.column_letter:3s} | {cell.value}")
    
    # 读取前10行数据
    print("\n" + "=" * 80)
    print("前10行数据:")
    print("=" * 80)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(11, ws.max_row), values_only=True), 1):
        print(f"\n--- 行 {i+1} ---")
        for j, value in enumerate(row):
            if j < len(headers):
                print(f"  {headers[j]}: {value}")
    
    # 统计非空行数
    non_empty_rows = 0
    for row in ws.iter_rows(min_row=2):
        if any(cell.value for cell in row):
            non_empty_rows += 1
    
    print("\n" + "=" * 80)
    print(f"数据统计: 共 {non_empty_rows} 行数据（不含标题）")
    print("=" * 80)
    
except ImportError:
    print("错误: 需要安装 openpyxl")
    print("请运行: pip install openpyxl")
except Exception as e:
    print(f"错误: {e}")
    import traceback
    traceback.print_exc()

